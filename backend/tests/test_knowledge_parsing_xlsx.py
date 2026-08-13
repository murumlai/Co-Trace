"""Tests for XLSX RFC workbook discovery, categorization, and parsing."""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.knowledge import parsing


def _make_rfc_xlsx(path, rows, sheet_name="RFC Table", header=None):
    """Create a minimal RFC workbook at *path* with the given rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    if header is None:
        header = ["Failed Test Name / Error Name", "Error Message/ Bin / Issue / Findings", "RFCs"]
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(str(path))
    return str(path)


class TestRfcCategoryDetection:
    @pytest.mark.parametrize(
        "filename, expected",
        [
            ("N32828_RFC.xlsx", "rfc_knowledge"),
            ("N32828-201_RFC.xlsx", "rfc_knowledge"),
            ("N32828-201_rfc_data.xlsx", "rfc_knowledge"),
            # XLSX with HLD keyword still classified as hld (rfc keyword absent)
            ("N32828-201_HLD.xlsx", "hld"),
            # PDF with 'rfc' in name but no debug/learning keywords -> uncategorized
            ("N32828_RFC.pdf", "uncategorized"),
        ],
    )
    def test_detect_category(self, filename, expected):
        assert parsing.detect_category(filename) == expected


class TestProductFamilyCodeExtraction:
    @pytest.mark.parametrize(
        "filename, expected",
        [
            ("N32828_RFC.xlsx", "N32828"),
            ("N32828-201_RFC.xlsx", "N32828"),
            ("N32828-101_RFC.xlsx", "N32828"),
            ("M79060-001_Debug.pdf", "M79060"),
            ("random_notes.pdf", None),
        ],
    )
    def test_extract_product_family_code(self, filename, expected):
        assert parsing.extract_product_family_code(filename) == expected


class TestTempFileSkipping:
    def test_scan_skips_office_temp_files(self, tmp_path):
        src = tmp_path / "Product_Docs"
        src.mkdir()
        # Normal file
        good = src / "N32828_RFC.xlsx"
        good.write_bytes(b"fake")
        # Office temp/lock file
        temp = src / "~$N32828_RFC.xlsx"
        temp.write_bytes(b"fake")

        docs = parsing.scan_source_documents(
            source_dirs=[str(src)], globs=["*.xlsx"]
        )
        filenames = [d.filename for d in docs]
        assert "N32828_RFC.xlsx" in filenames
        assert "~$N32828_RFC.xlsx" not in filenames


class TestRfcWorkbookValidation:
    def test_rfc_without_product_code_raises(self, tmp_path):
        path = _make_rfc_xlsx(
            tmp_path / "plain_RFC.xlsx",
            [["TEST_FOO", "bad reading", "RFC-001"]],
        )
        # describe_document should warn
        doc = parsing.describe_document(path)
        assert doc.category == "rfc_knowledge"
        assert any("no product code" in w.lower() or "product-family" in w for w in doc.warnings)
        # parse_document should raise
        with pytest.raises(ValueError, match="product code"):
            parsing.parse_document(doc)

    def test_rfc_with_no_extractable_rows_raises(self, tmp_path):
        path = _make_rfc_xlsx(
            tmp_path / "N32828_RFC.xlsx",
            [],  # header only, no data rows
        )
        doc = parsing.describe_document(path)
        with pytest.raises(ValueError, match="No RFC rows"):
            parsing.parse_document(doc)


class TestXlsxRfcParsing:
    def test_basic_row_becomes_section(self, tmp_path):
        path = _make_rfc_xlsx(
            tmp_path / "N32828_RFC.xlsx",
            [["POWER_TEST_01", "12V droop at high load", "RFC-1234"]],
        )
        doc = parsing.describe_document(path)
        assert doc.product_code is None  # no revision suffix
        assert doc.product_family_code == "N32828"
        assert doc.category == "rfc_knowledge"

        content_hash, sections = parsing.parse_document(doc)
        assert content_hash
        assert len(sections) == 1
        s = sections[0]
        assert s.category == "rfc_knowledge"
        assert "POWER_TEST_01" in s.heading
        assert "POWER_TEST_01" in s.text
        assert "RFC-1234" in s.text
        assert "12V droop" in s.text

    def test_revisioned_filename_sets_both_codes(self, tmp_path):
        path = _make_rfc_xlsx(
            tmp_path / "N32828-201_RFC.xlsx",
            [["USB_ENUM_FAIL", "USB not detected", "RFC-0001"]],
        )
        doc = parsing.describe_document(path)
        assert doc.product_code == "N32828-201"
        assert doc.product_family_code == "N32828"

    def test_empty_rows_are_skipped(self, tmp_path):
        path = _make_rfc_xlsx(
            tmp_path / "N32828_RFC.xlsx",
            [
                ["GOOD_TEST", "some error", "RFC-001"],
                ["", "", ""],        # fully empty
                ["", "only msg", "RFC-002"],  # no test name -> skip
                [None, None, None],
            ],
        )
        _, sections = parsing.parse_document(parsing.describe_document(path))
        assert len(sections) == 1
        assert "GOOD_TEST" in sections[0].heading

    def test_multi_rfc_cell_is_split(self, tmp_path):
        path = _make_rfc_xlsx(
            tmp_path / "N32828_RFC.xlsx",
            [["MULTI_RFC_TEST", "multiple failures", "RFC-001, RFC-002; RFC-003"]],
        )
        _, sections = parsing.parse_document(parsing.describe_document(path))
        assert len(sections) == 1
        text = sections[0].text
        assert "RFC-001" in text
        assert "RFC-002" in text
        assert "RFC-003" in text

    def test_multi_rfc_newline_separator(self, tmp_path):
        path = _make_rfc_xlsx(
            tmp_path / "N32828_RFC.xlsx",
            [["NL_TEST", "error", "RFC-A\nRFC-B\nRFC-C"]],
        )
        _, sections = parsing.parse_document(parsing.describe_document(path))
        text = sections[0].text
        assert "RFC-A" in text and "RFC-B" in text and "RFC-C" in text

    def test_all_sheets_processed(self, tmp_path):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Failed Test Name / Error Name", "Error Message/ Bin / Issue / Findings", "RFCs"])
        ws1.append(["TEST_A", "error a", "RFC-100"])

        ws2 = wb.create_sheet(title="Sheet2")
        ws2.append(["Failed Test Name / Error Name", "Error Message/ Bin / Issue / Findings", "RFCs"])
        ws2.append(["TEST_B", "error b", "RFC-200"])

        path = str(tmp_path / "N32828_RFC.xlsx")
        wb.save(path)

        _, sections = parsing.parse_document(parsing.describe_document(path))
        headings = [s.heading for s in sections]
        assert any("TEST_A" in h for h in headings)
        assert any("TEST_B" in h for h in headings)

    def test_content_hash_is_stable(self, tmp_path):
        rows = [["STABLE_TEST", "stable error", "RFC-999"]]
        path = str(tmp_path / "N32828_RFC.xlsx")
        _make_rfc_xlsx(path, rows)
        doc = parsing.describe_document(path)
        h1, _ = parsing.parse_document(doc)
        h2, _ = parsing.parse_document(doc)
        assert h1 == h2

    def test_header_row_case_insensitive(self, tmp_path):
        # Use differently-cased header
        path = _make_rfc_xlsx(
            tmp_path / "N32828_RFC.xlsx",
            [["CASE_TEST", "err", "RFC-777"]],
            header=["failed test name / error name", "error message/ bin / issue / findings", "rfcs"],
        )
        _, sections = parsing.parse_document(parsing.describe_document(path))
        assert len(sections) == 1
        assert "CASE_TEST" in sections[0].heading
